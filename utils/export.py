from flask import make_response

from datetime import datetime
from dateutil import parser

from decimal import Decimal

import io
from openpyxl import Workbook
from openpyxl.styles import Alignment, numbers, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter
from collections import defaultdict

from flask_app import app
from utils.schema import db, Cobranzas, Precios, LiquidacionViajes, LiquidacionGastos
from utils.utils import redondear


def exportar_cobranza(fecha_creacion):
    cobranzas = Cobranzas.query.filter_by(fecha_creacion=fecha_creacion).all()

    # Crea un diccionario para almacenar las sumas de subtotales por grupo
    subtotales_origen = defaultdict(int)
    subtotales_destino = defaultdict(int)
    subtotales_dif = defaultdict(int)
    subtotales_tolerancia = defaultdict(int)
    subtotales_dif_tol = defaultdict(int)
    subtotales = defaultdict(int)

    cambio_de_grupo = defaultdict(int)
    for cobranza in cobranzas:
        diferencia = cobranza.kilos_destino - cobranza.kilos_origen
        tolerancia = redondear(Decimal(cobranza.kilos_destino) * Decimal(0.002))

        grupo = (cobranza.origen, cobranza.destino)
        subtotales_origen[grupo] += cobranza.kilos_origen
        subtotales_destino[grupo] += cobranza.kilos_destino
        subtotales_dif[grupo] += diferencia
        subtotales_tolerancia[grupo] += tolerancia
        subtotales_dif_tol[grupo] += diferencia + tolerancia
        subtotales[grupo] += redondear(Decimal(cobranza.precio) * Decimal(cobranza.kilos_destino))

        cambio_de_grupo[grupo] = cobranza.tiquet

    # Ordena las cobranzas por las columnas 'origen' y 'destino'
    cobranzas_ordenadas = sorted(cobranzas, key=lambda cobranza: (cobranza.origen, cobranza.destino, cobranza.fecha_viaje, cobranza.chofer))

    # Crear un archivo Excel en memoria
    output = io.BytesIO()
    workbook = Workbook()
    sheet = workbook.active

    sheet.column_dimensions['A'].width = 2.64
    sheet.column_dimensions['B'].width = 11.00
    sheet.column_dimensions['C'].width = 20.55
    sheet.column_dimensions['D'].width = 9.09
    sheet.column_dimensions['E'].width = 11.82
    sheet.column_dimensions['F'].width = 18.64
    sheet.column_dimensions['G'].width = 17.64
    sheet.column_dimensions['H'].width = 9.91
    sheet.column_dimensions['I'].width = 10.91
    sheet.column_dimensions['J'].width = 11.09
    sheet.column_dimensions['K'].width = 7.18
    sheet.column_dimensions['L'].width = 6.27
    sheet.column_dimensions['M'].width = 6.36
    sheet.column_dimensions['N'].width = 6.27
    sheet.column_dimensions['O'].width = 14.64

    # Agregar la fecha como la primera fila
    sheet.append([])  # Agregar una fila en blanco después de la fecha
    sheet.append(['D & R TRANSPORTES'])  # Agregar una fila en blanco después de la fecha

        # Obtener el rango de columnas con valores None
    inicio_columna = 1  # Cambiar al índice de la primera columna con valor None
    fin_columna = 15   # Cambiar al índice de la última columna con valor None

    # Combinar las celdas en el rango de columnas
    sheet.merge_cells(start_row=sheet.max_row, start_column=inicio_columna, end_row=sheet.max_row, end_column=fin_columna)

    # Centrar el contenido en la celda combinada
    merged_cell = sheet.cell(row=sheet.max_row, column=inicio_columna)
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Aplicar el estilo de fuente deseado (Arial Black, size 22, purple color)
    font = Font(name='Arial Black', size=22, color="800080")  # Using a standard purple color index
    merged_cell.font = font

    sheet.row_dimensions[2].height = 35

    sheet.append([])  # Agregar una fila en blanco después de la fecha
    sheet.append([None, datetime.now().strftime('%d/%m/%Y')])
    sheet.append([])  # Agregar una fila en blanco después de la fecha

    # Agregar encabezados
    encabezados = ['N°', 'Fecha', 'Chofer', 'Chapa', 'Producto', 'Origen', 'Destino', 'Tiquet', 'Kilos Origen', 'Kilos Destino', 'Dif.', 'Tolera', 'Dif. Tol.', 'Precio', 'Total']
    sheet.append(encabezados)


    # Aplicar bordes y relleno a las celdas del encabezado
    for col_idx, header_text in enumerate(encabezados, start=1):
        col_letter = get_column_letter(col_idx)
        cell = sheet[f'{col_letter}6']
        
        # Aplicar bordes
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.border = thin_border
        
        # Aplicar relleno con el color Gold, Accent 4, Lighter 40%
        fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        cell.fill = fill

    # Aplicar alineación vertical y horizontal en la celda
        cell.alignment = Alignment(horizontal='left', vertical='bottom')


    sheet.row_dimensions[6].height = 30

    # Agregar filas de datos
    contador = 1
    for cobranza in cobranzas_ordenadas:
        diferencia = cobranza.kilos_destino - cobranza.kilos_origen
        tolerancia = redondear(Decimal(cobranza.kilos_origen) * Decimal(0.002))

        fila = [contador, 
                cobranza.fecha_creacion.strftime('%d/%m/%Y'), 
                cobranza.chofer, 
                cobranza.chapa,
                cobranza.producto,
                cobranza.origen, 
                cobranza.destino, 
                cobranza.tiquet, 
                cobranza.kilos_origen, 
                cobranza.kilos_destino,
                diferencia, 
                tolerancia,  
                diferencia + tolerancia,
                cobranza.precio, 
                redondear(Decimal(cobranza.precio) * Decimal(cobranza.kilos_destino))]

        sheet.append(fila)

        for col in range(8, 16):
            cell = sheet.cell(row=sheet.max_row, column=col)

            if col == 14:
                cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED2
            else:
                cell.number_format = "#,##0"
        
        for col in range(1, 16):
            cell = sheet.cell(row=sheet.max_row, column=col)
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            cell.border = thin_border

        contador += 1

        grupo = (cobranza.origen, cobranza.destino)

        if cambio_de_grupo[grupo] == cobranza.tiquet:
            sheet.append(['Subtotal', None, None, None, None, None, None, None, 
                        subtotales_origen[grupo],
                        subtotales_destino[grupo],
                        subtotales_dif[grupo], 
                        subtotales_tolerancia[grupo],
                        subtotales_dif_tol[grupo], 
                        None, subtotales[grupo]])
            
            # Obtener el rango de columnas con valores None
            inicio_columna = 1  # Cambiar al índice de la primera columna con valor None
            fin_columna = 8   # Cambiar al índice de la última columna con valor None

            # Combinar las celdas en el rango de columnas
            sheet.merge_cells(start_row=sheet.max_row, start_column=inicio_columna, end_row=sheet.max_row, end_column=fin_columna)

            # Centrar el contenido en la celda combinada
            merged_cell = sheet.cell(row=sheet.max_row, column=inicio_columna)
            merged_cell.alignment = Alignment(horizontal='center', vertical='center')

            # Formatear columnas 8 a 15 como números
            for col in range(8, 16):
                cell = sheet.cell(row=sheet.max_row, column=col)

                if col == 14:
                    cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED2
                else:
                    cell.number_format = "#,##0"

            for col in range(1, 16):
                cell = sheet.cell(row=sheet.max_row, column=col)
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                cell.border = thin_border

                # Aplicar relleno con el color Gray, Accent 4, Lighter 60%
                fill = PatternFill(start_color="969696", end_color="969696", fill_type="solid")  # Gray, Accent 4, Lighter 60%
                cell.fill = fill

    total_origen = 0
    total_destino = 0
    total_dif = 0
    total_tolerancia = 0
    total_dif_tol = 0
    total_gs = 0

    for grupo in subtotales_origen:
        total_origen += subtotales_origen[grupo]

    for grupo in subtotales_destino:
        total_destino += subtotales_destino[grupo]

    for grupo in subtotales_dif:
        total_dif += subtotales_dif[grupo]

    for grupo in subtotales_tolerancia:
        total_tolerancia += subtotales_tolerancia[grupo]

    for grupo in subtotales_dif_tol:
        total_dif_tol += subtotales_dif_tol[grupo]

    for grupo in subtotales:
        total_gs += subtotales[grupo]

    sheet.append(['TOTAL', None, None, None, None, None, None, None, 
            total_origen,
            total_destino,
            total_dif, 
            total_tolerancia,
            total_dif_tol, 
            None, total_gs])
        
    # Obtener el rango de columnas con valores None
    inicio_columna = 1  # Cambiar al índice de la primera columna con valor None
    fin_columna = 8   # Cambiar al índice de la última columna con valor None

    # Combinar las celdas en el rango de columnas
    sheet.merge_cells(start_row=sheet.max_row, start_column=inicio_columna, end_row=sheet.max_row, end_column=fin_columna)

    # Centrar el contenido en la celda combinada
    merged_cell = sheet.cell(row=sheet.max_row, column=inicio_columna)
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Formatear columnas 8 a 15 como números
    for col in range(8, 16):
        cell = sheet.cell(row=sheet.max_row, column=col)

        if col == 14:
            cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED2
        else:
            cell.number_format = "#,##0"

    for col in range(1, 16):
        cell = sheet.cell(row=sheet.max_row, column=col)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.border = thin_border

        # Aplicar relleno con el color Gray, Accent 4, Lighter 60%
        fill = PatternFill(start_color="969696", end_color="969696", fill_type="solid")  # Gray, Accent 4, Lighter 60%
        cell.fill = fill

    # Guardar el archivo Excel en el flujo de salida
    workbook.save(output)
    output.seek(0)

    # Crear la respuesta para el cliente con el archivo Excel
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=cobranza_{fecha_creacion}.xlsx'
    app.logger.warning(f'Cobranza exportada {fecha_creacion}')

    return response


def exportar_informe(fecha_inicio, fecha_fin):
    fecha_inicio_parsed = parser.isoparse(fecha_inicio).date()
    fecha_fin_parsed = parser.isoparse(fecha_fin).date()

    # Query the database for records within the date range
    cobranzas = Cobranzas.query.filter(Cobranzas.fecha_creacion.between(fecha_inicio_parsed, fecha_fin_parsed)).all()

    # Ordena las cobranzas por las columnas 'origen' y 'destino'
    cobranzas_ordenadas = sorted(cobranzas, key=lambda cobranza: cobranza.chapa)

    # Crear un archivo Excel en memoria
    output = io.BytesIO()
    workbook = Workbook()
    sheet = workbook.active

    sheet.column_dimensions['A'].width = 2.64
    sheet.column_dimensions['B'].width = 11.00
    sheet.column_dimensions['C'].width = 20.55
    sheet.column_dimensions['D'].width = 9.09
    sheet.column_dimensions['E'].width = 11.82
    sheet.column_dimensions['F'].width = 18.64
    sheet.column_dimensions['G'].width = 17.64
    sheet.column_dimensions['H'].width = 9.91
    sheet.column_dimensions['I'].width = 10.91
    sheet.column_dimensions['J'].width = 11.09
    sheet.column_dimensions['K'].width = 7.18
    sheet.column_dimensions['L'].width = 6.27
    sheet.column_dimensions['M'].width = 6.36
    sheet.column_dimensions['N'].width = 6.27
    sheet.column_dimensions['O'].width = 14.64

    # Agregar encabezados
    encabezados = ['N°', 'Fecha', 'Chofer', 'Chapa', 'Producto', 'Origen', 'Destino', 'Tiquet', 'Kilos Origen', 'Kilos Destino', 'Dif.', 'Tolera', 'Dif. Tol.', 'Precio', 'Total']
    sheet.append(encabezados)

    # Congelar la fila del encabezado
    sheet.freeze_panes = 'A2'

    # Agregar filas de datos
    contador = 1
    for cobranza in cobranzas_ordenadas:
        diferencia = cobranza.kilos_destino - cobranza.kilos_origen
        tolerancia = redondear(Decimal(cobranza.kilos_origen) * Decimal(0.002))

        fila = [contador, 
                cobranza.fecha_viaje.strftime('%d/%m/%Y'), 
                cobranza.chofer, 
                cobranza.chapa,
                cobranza.producto,
                cobranza.origen, 
                cobranza.destino, 
                cobranza.tiquet, 
                cobranza.kilos_origen, 
                cobranza.kilos_destino,
                diferencia, 
                tolerancia,  
                diferencia + tolerancia,
                cobranza.precio, 
                redondear(Decimal(cobranza.precio) * Decimal(cobranza.kilos_destino))]
        
        
        sheet.append(fila)

        for col in range(8, 16):
            cell = sheet.cell(row=sheet.max_row, column=col)

            if col == 14:
                cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED2
            else:
                cell.number_format = "#,##0"

        contador += 1

    # Guardar el archivo Excel en el flujo de salida
    workbook.save(output)
    output.seek(0)

    # Crear la respuesta para el cliente con el archivo Excel
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=informe_{fecha_inicio}-{fecha_fin}.xlsx'
    app.logger.warning(f'Informe exportado: {fecha_inicio}-{fecha_fin}')

    return response


def exportar_liquidacion(chofer, fecha):
    # Perform an inner join between Cobranza and Liquidaciones based on 'chofer' and 'fecha_de_liquidacion'
    viajes = db.session.query(Cobranzas, LiquidacionViajes).join(
        LiquidacionViajes,
        Cobranzas.id == LiquidacionViajes.id
    ).filter(
        Cobranzas.chofer == chofer,
        LiquidacionViajes.fecha_liquidacion == fecha
    ).all()

    # Perform an inner join between Cobranza and Liquidaciones based on 'chofer' and 'fecha_de_liquidacion'
    gastos_sin_boleta = LiquidacionGastos.query.filter_by(chofer=chofer, fecha_liquidacion=fecha, boleta=None).all()
    gastos_con_boleta = LiquidacionGastos.query.filter_by(chofer=chofer, fecha_liquidacion=fecha).filter(LiquidacionGastos.boleta.isnot(None)).all()

    # Calcular la longitud máxima de las tres listas
    max_len = max(len(viajes), len(gastos_sin_boleta), len(gastos_con_boleta))

    # Rellenar las listas para que tengan la misma longitud con None si es necesario
    viajes += [None] * (max_len - len(viajes))
    gastos_sin_boleta += [None] * (max_len - len(gastos_sin_boleta))
    gastos_con_boleta += [None] * (max_len - len(gastos_con_boleta))

    # Combinar las tres listas en una lista de tuplas usando zip
    results = zip(viajes, gastos_sin_boleta, gastos_con_boleta)

    # Crear un archivo Excel en memoria
    output = io.BytesIO()
    workbook = Workbook()
    sheet = workbook.active

    sheet.column_dimensions['A'].width = 2.64
    sheet.column_dimensions['B'].width = 9.91
    sheet.column_dimensions['C'].width = 7.91
    sheet.column_dimensions['D'].width = 9.36
    sheet.column_dimensions['E'].width = 14.82
    sheet.column_dimensions['F'].width = 11.36
    sheet.column_dimensions['G'].width = 10.3
    sheet.column_dimensions['H'].width = 11.0
    sheet.column_dimensions['I'].width = 8.09
    sheet.column_dimensions['J'].width = 9.60
    sheet.column_dimensions['K'].width = 10.27
    sheet.column_dimensions['L'].width = 10.82
    sheet.column_dimensions['M'].width = 10.27
    sheet.column_dimensions['N'].width = 10.36
    sheet.column_dimensions['O'].width = 8.91
    sheet.column_dimensions['P'].width = 10.82

    # Agregar la fecha como la primera fila
    sheet.append([])  # Agregar una fila en blanco después de la fecha
    sheet.append(['LIQUIDACION DE FLETES'])  # Agregar una fila en blanco después de la fecha

        # Obtener el rango de columnas con valores None
    inicio_columna = 1  # Cambiar al índice de la primera columna con valor None
    fin_columna = 16   # Cambiar al índice de la última columna con valor None

    # Combinar las celdas en el rango de columnas
    sheet.merge_cells(start_row=sheet.max_row, start_column=inicio_columna, end_row=sheet.max_row, end_column=fin_columna)

    # Centrar el contenido en la celda combinada
    merged_cell = sheet.cell(row=sheet.max_row, column=inicio_columna)
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Aplicar bordes
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for col in range(1, 17):
        cell = sheet.cell(row=sheet.max_row, column=col)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.border = thin_border
        cell.font = Font(bold=True) 

    sheet.row_dimensions[2].height = 21

    if viajes and viajes[0]:
        chapa = viajes[0][0].chapa
    else:
        chapa = ""

    sheet.append([f'Conductor: {chofer}                Chapa: {chapa}                Fecha: {datetime.now().strftime("%d/%m/%Y")}'])  # Agregar una fila en blanco después de la fecha
            # Obtener el rango de columnas con valores None
    inicio_columna = 1  # Cambiar al índice de la primera columna con valor None
    fin_columna = 16   # Cambiar al índice de la última columna con valor None

    # Combinar las celdas en el rango de columnas
    sheet.merge_cells(start_row=sheet.max_row, start_column=inicio_columna, end_row=sheet.max_row, end_column=fin_columna)

    # Centrar el contenido en la celda combinada
    merged_cell = sheet.cell(row=sheet.max_row, column=inicio_columna)
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for col in range(1, 17):
        cell = sheet.cell(row=sheet.max_row, column=col)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.border = thin_border
        cell.font = Font(bold=True) 

    sheet.row_dimensions[3].height = 21

    sheet.append(['FLETES', None, None, None, None, None, None, None, None, None, None,
            'GASTOS (VIATICO/GASOIL)', None, None, None, None])


    # Obtener el rango de columnas con valores None
    inicio_columna = 1  # Cambiar al índice de la primera columna con valor None
    fin_columna = 11   # Cambiar al índice de la última columna con valor None

    # Combinar las celdas en el rango de columnas
    sheet.merge_cells(start_row=sheet.max_row, start_column=inicio_columna, end_row=sheet.max_row, end_column=fin_columna)

    # Centrar el contenido en la celda combinada
    merged_cell = sheet.cell(row=sheet.max_row, column=inicio_columna)
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for col in range(1, 17):
        cell = sheet.cell(row=sheet.max_row, column=col)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.border = thin_border
        cell.font = Font(bold=True) 

    # Obtener el rango de columnas con valores None
    inicio_columna = 12  # Cambiar al índice de la primera columna con valor None
    fin_columna = 16   # Cambiar al índice de la última columna con valor None

    # Combinar las celdas en el rango de columnas
    sheet.merge_cells(start_row=sheet.max_row, start_column=inicio_columna, end_row=sheet.max_row, end_column=fin_columna)

    # Centrar el contenido en la celda combinada
    merged_cell = sheet.cell(row=sheet.max_row, column=inicio_columna)
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for col in range(1, 17):
        cell = sheet.cell(row=sheet.max_row, column=col)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.border = thin_border
        cell.font = Font(bold=True) 

    # Agregar encabezados
    encabezados = ['N°', 'Fecha', 'Producto', 'Recepcion N°', 'Origen', 'Destino', 'Kilos Origen', 'Kilos Llegada', 
                   'Dif.', 'Gs. p/ KILO', 'Importe Gs.', 'Fecha', 'Importe Gs.', 'Fecha', 'Boleta N°', 'Importe Gs.']
    sheet.append(encabezados)

    for col in range(1, 17):
        cell = sheet.cell(row=sheet.max_row, column=col)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.border = thin_border
        cell.font = Font(bold=True) 

    sheet.row_dimensions[5].height = 25

    # Agregar filas de datos
    subtotal_viajes = 0
    subtotal_sin_boleta = 0
    subtotal_con_boleta = 0

    contador = 1
    for viaje, sin_boleta, con_boleta in results:
        fila = [contador]

        if viaje:
            viaje_fila = [viaje[0].fecha_viaje.strftime('%d/%m/%Y'), viaje[0].producto, 
                          viaje[0].tiquet, viaje[0].origen, viaje[0].destino, 
                          viaje[0].kilos_origen, viaje[0].kilos_destino, 
                           viaje[0].kilos_destino - viaje[0].kilos_origen, viaje[1].precio_liquidacion, 
                          redondear(Decimal(viaje[1].precio_liquidacion) * Decimal(viaje[0].kilos_destino))]
            fila.extend(viaje_fila)
            subtotal_viajes += redondear(Decimal(viaje[1].precio_liquidacion) * Decimal(viaje[0].kilos_destino))
        else:
            fila.extend([None] * 10)

        if sin_boleta:
            sin_boleta_fila = [sin_boleta.fecha.strftime('%d/%m/%Y'), sin_boleta.importe]
            subtotal_sin_boleta += sin_boleta.importe
            fila.extend(sin_boleta_fila)
        else:
            fila.extend([None] * 2)


        if con_boleta:
            con_boleta_fila = [con_boleta.fecha.strftime('%d/%m/%Y'), con_boleta.boleta, con_boleta.importe]
            subtotal_con_boleta += con_boleta.importe
            fila.extend(con_boleta_fila)
        else:
            fila.extend([None] * 3)
        
        sheet.append(fila)

        for col in range(1, 17):
            cell = sheet.cell(row=sheet.max_row, column=col)

            if col == 10:
                cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED2
            else:
                cell.number_format = "#,##0"

            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            cell.border = thin_border

        contador += 1

    subtotales = [None, None, None, None, None, None, None, None, None, None, None, 'Subtotal', subtotal_sin_boleta, None, 'Subtotal', subtotal_con_boleta]
    sheet.append(subtotales)

    for col in range(1, 17):
        cell = sheet.cell(row=sheet.max_row, column=col)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.border = thin_border
        cell.font = Font(bold=True) 

    for col in [13, 16]:
        cell = sheet.cell(row=sheet.max_row, column=col)
        cell.number_format = "#,##0"
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.border = thin_border

    total_gastos = subtotal_sin_boleta + subtotal_con_boleta
    total = [None, None, None, None, None, None, None, None, 'TOTAL FLETES:', None, subtotal_viajes, 'TOTAL GASTOS:', None, None, None, total_gastos]
    sheet.append(total)

    for col in [11, 16]:
        cell = sheet.cell(row=sheet.max_row, column=col)
        cell.number_format = "#,##0"
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.border = thin_border

    for col in range(1, 17):
        cell = sheet.cell(row=sheet.max_row, column=col)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.border = thin_border
        cell.font = Font(bold=True) 

    sheet.append([None])

    total_cobrar = [None, None, None, None, None, None, None, 'TOTAL A COBRAR:', None, None, subtotal_viajes - total_gastos]
    sheet.append(total_cobrar)
    
    for col in range(8, 12):
        cell = sheet.cell(row=sheet.max_row, column=col)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.border = thin_border
        cell.font = Font(bold=True) 

    total_facturar = [None, None, None, None, None, None, None, 'TOTAL A FACTURAR:', None, None, subtotal_viajes - subtotal_con_boleta]
    sheet.append(total_facturar)

    for col in range(8, 12):
        cell = sheet.cell(row=sheet.max_row, column=col)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.border = thin_border
        cell.font = Font(bold=True)
    

    # Guardar el archivo Excel en el flujo de salida
    workbook.save(output)
    output.seek(0)

    # Crear la respuesta para el cliente con el archivo Excel
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={chofer}_Liquidacion_{fecha}.xlsx'
    app.logger.warning(f'Liquidacion {chofer}_{fecha} exportada')
    
    return response


def exportar_precios():
    precios = Precios.query.all()

    # Ordena las cobranzas por las columnas 'origen' y 'destino'
    precios_ordenados = sorted(precios, key=lambda precios: (precios.origen, precios.destino))

    # Crear un archivo Excel en memoria
    output = io.BytesIO()
    workbook = Workbook()
    sheet = workbook.active

    # Agregar encabezados
    encabezados = ['Origen', 'Destino', 'Precio', 'Precio de Liquidación']
    sheet.append(encabezados)

    # Ajustar el ancho de las columnas
    for col_idx in range(1, 5):
        column_letter = get_column_letter(col_idx)
        sheet.column_dimensions[column_letter].width = 20

    # Estilo de borde
    border_style = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))

    # Aplicar el estilo de borde a cada celda en la fila
    for cell in sheet[sheet.max_row]:
        cell.border = border_style

    # Agregar filas de datos
    for precio in precios_ordenados:
        fila = [precio.origen, precio.destino, precio.precio, precio.precio_liquidacion]
        
        sheet.append(fila)

        sheet.cell(row=sheet.max_row, column=3).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
        sheet.cell(row=sheet.max_row, column=4).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

        # Aplicar el estilo de borde a cada celda en la fila
        for cell in sheet[sheet.max_row]:
            cell.border = border_style

    # Guardar el archivo Excel en el flujo de salida
    workbook.save(output)
    output.seek(0)

    # Crear la respuesta para el cliente con el archivo Excel
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=lista_de_precios.xlsx'
    
    app.logger.warning("Lista de Precios exportada")
    return response