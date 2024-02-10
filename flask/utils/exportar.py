from flask import make_response

from datetime import datetime
from dateutil import parser

import io
from openpyxl import Workbook
from openpyxl.styles import Alignment, numbers, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter

from app_database import logger
from utils.schema import db, Cobranzas, Precios, LiquidacionViajes, LiquidacionGastos, Liquidaciones


def exportar_cobranza(fecha_creacion):
    cobranzas_ordenadas = Cobranzas.query.filter_by(fecha_creacion=fecha_creacion).order_by(Cobranzas.producto, Cobranzas.origen, Cobranzas.destino, Cobranzas.chofer, Cobranzas.fecha_viaje).all()
    # Crea un diccionario para almacenar las sumas de subtotales por grupo

    subtotales_grupo = {}
    default_grupo = {'origen': 0, 'destino': 0, 'diferencia': '', 'tolerancia': '', 'last_row': 0, 
                     'diferencia_tolerancia': '', 'subtotal': '', 'ultima_entrada': '', 'producto': ''}

    group_counter = 6
    first_row = 0
    for cobranza in cobranzas_ordenadas:
        grupo = (cobranza.origen, cobranza.destino, cobranza.producto)

        if grupo not in subtotales_grupo:
            group_counter += 1
            subtotales_grupo[grupo] = default_grupo.copy()
            first_row = group_counter
            subtotales_grupo[grupo]['producto'] = cobranza.producto

        subtotales_grupo[grupo]['origen'] += cobranza.kilos_origen
        subtotales_grupo[grupo]['destino'] += cobranza.kilos_destino
        subtotales_grupo[grupo]['diferencia'] = f'=SUM(K{first_row}:K{group_counter})'
        subtotales_grupo[grupo]['tolerancia'] = f'=SUM(L{first_row}:L{group_counter})'
        subtotales_grupo[grupo]['diferencia_tolerancia'] = f'=SUM(M{first_row}:M{group_counter})'
        subtotales_grupo[grupo]['subtotal'] = f'=SUM(O{first_row}:O{group_counter})'

        subtotales_grupo[grupo]['ultima_entrada'] = cobranza.tiquet
        subtotales_grupo[grupo]['last_row'] = group_counter

        group_counter += 1

    # Crear un archivo Excel en memoria
    output = io.BytesIO()
    workbook = Workbook()
    sheet = workbook.active

    sheet.column_dimensions['A'].width = 2.64
    sheet.column_dimensions['B'].width = 11.00
    sheet.column_dimensions['C'].width = 20.55
    sheet.column_dimensions['D'].width = 9.09
    sheet.column_dimensions['E'].width = 11.82
    sheet.column_dimensions['F'].width = 18.64
    sheet.column_dimensions['G'].width = 17.64
    sheet.column_dimensions['H'].width = 9.91
    sheet.column_dimensions['I'].width = 10.91
    sheet.column_dimensions['J'].width = 11.09
    sheet.column_dimensions['K'].width = 7.18
    sheet.column_dimensions['L'].width = 6.27
    sheet.column_dimensions['M'].width = 6.36
    sheet.column_dimensions['N'].width = 6.27
    sheet.column_dimensions['O'].width = 14.64

    # Agregar la fecha como la primera fila
    sheet.append([])  # Agregar una fila en blanco después de la fecha
    # Agregar una fila en blanco después de la fecha
    sheet.append(['D & R TRANSPORTES'])

    # Obtener el rango de columnas con valores None
    inicio_columna = 1  # Cambiar al índice de la primera columna con valor None
    fin_columna = 15   # Cambiar al índice de la última columna con valor None

    # Combinar las celdas en el rango de columnas
    sheet.merge_cells(start_row=sheet.max_row, start_column=inicio_columna,
                      end_row=sheet.max_row, end_column=fin_columna)

    # Centrar el contenido en la celda combinada
    merged_cell = sheet.cell(row=sheet.max_row, column=inicio_columna)
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Aplicar el estilo de fuente deseado (Arial Black, size 22, purple color)
    # Using a standard purple color index
    font = Font(name='Arial Black', size=22, color="800080")
    merged_cell.font = font

    sheet.row_dimensions[2].height = 35

    sheet.append([])  # Agregar una fila en blanco después de la fecha
    sheet.append([None, datetime.now().strftime('%d/%m/%Y')])
    sheet.append([])  # Agregar una fila en blanco después de la fecha

    # Agregar encabezados
    encabezados = ['N°', 'Fecha', 'Chofer', 'Chapa', 'Producto', 'Origen', 'Destino', 'Tiquet',
                   'Kilos Origen', 'Kilos Destino', 'Dif.', 'Tolera', 'Dif. Tol.', 'Precio', 'Total']
    sheet.append(encabezados)

    # Aplicar bordes y relleno a las celdas del encabezado
    for col_idx, _ in enumerate(encabezados, start=1):
        col_letter = get_column_letter(col_idx)
        cell = sheet[f'{col_letter}6']

        # Aplicar bordes
        thin_border = Border(left=Side(style='thin'), right=Side(
            style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.border = thin_border

        # Aplicar relleno con el color Gold, Accent 4, Lighter 40%
        fill = PatternFill(start_color="FFC000",
                           end_color="FFC000", fill_type="solid")
        cell.fill = fill

    # Aplicar alineación vertical y horizontal en la celda
        cell.alignment = Alignment(horizontal='left', vertical='bottom')

    sheet.row_dimensions[6].height = 30

    # Agregar filas de datos
    index = 1
    contador = 7
    for cobranza in cobranzas_ordenadas:
        fila = [index,
                cobranza.fecha_viaje.strftime('%d/%m/%Y'),
                cobranza.chofer,
                cobranza.chapa,
                cobranza.producto,
                cobranza.origen,
                cobranza.destino,
                cobranza.tiquet,
                cobranza.kilos_origen,
                cobranza.kilos_destino,
                f'=+J{contador}-I{contador}',
                f'=ROUND(J{contador}*0.002, 0)',
                f'=+L{contador}+K{contador}',
                cobranza.precio,
                f'=ROUND(J{contador}*N{contador}, 0)']

        sheet.append(fila)

        for col in range(8, 16):
            cell = sheet.cell(row=sheet.max_row, column=col)

            if col == 14:
                cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED2
            else:
                cell.number_format = "#,##0"

        for col in range(1, 16):
            cell = sheet.cell(row=sheet.max_row, column=col)
            thin_border = Border(left=Side(style='thin'), right=Side(
                style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            cell.border = thin_border

        grupo = (cobranza.origen, cobranza.destino, cobranza.producto)
        if subtotales_grupo[grupo]['ultima_entrada'] == cobranza.tiquet:
            contador += 1

            sheet.append(['Subtotal', None, None, None, None, None, None, None,
                          subtotales_grupo[grupo]['origen'],
                          subtotales_grupo[grupo]['destino'],
                          subtotales_grupo[grupo]['diferencia'],
                          subtotales_grupo[grupo]['tolerancia'],
                          subtotales_grupo[grupo]['diferencia_tolerancia'],
                          None, subtotales_grupo[grupo]['subtotal']
                          ])

            # Obtener el rango de columnas con valores None
            inicio_columna = 1  # Cambiar al índice de la primera columna con valor None
            fin_columna = 8   # Cambiar al índice de la última columna con valor None

            # Combinar las celdas en el rango de columnas
            sheet.merge_cells(start_row=sheet.max_row, start_column=inicio_columna,
                              end_row=sheet.max_row, end_column=fin_columna)

            # Centrar el contenido en la celda combinada
            merged_cell = sheet.cell(row=sheet.max_row, column=inicio_columna)
            merged_cell.alignment = Alignment(
                horizontal='center', vertical='center')

            # Formatear columnas 8 a 15 como números
            for col in range(8, 16):
                cell = sheet.cell(row=sheet.max_row, column=col)

                if col == 14:
                    cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED2
                else:
                    cell.number_format = "#,##0"

            for col in range(1, 16):
                cell = sheet.cell(row=sheet.max_row, column=col)
                thin_border = Border(left=Side(style='thin'), right=Side(
                    style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                cell.border = thin_border

                # Aplicar relleno con el color Gray, Accent 4, Lighter 60%
                # Gray, Accent 4, Lighter 60%
                fill = PatternFill(start_color="969696",
                                   end_color="969696", fill_type="solid")
                cell.fill = fill

        contador += 1
        index += 1

    total = {'origen': '=', 'destino': '=', 'diferencia': '=', 'tolerancia': '=', 'diferencia_tolerancia': '=', 'total': '=', 'productos': {}}
    last_row = None
    for grupo, subtotal_agrupado in subtotales_grupo.items():
        last_row = subtotal_agrupado["last_row"] + 1

        total['origen'] += f'+I{last_row}'
        total['destino'] += f'+J{last_row}'
        total['diferencia'] += f'+K{last_row}'
        total['tolerancia'] += f'+L{last_row}'
        total['diferencia_tolerancia'] += f'+M{last_row}'
        subtotal_row = f'+O{last_row}'
        total['total'] += subtotal_row

        producto = subtotal_agrupado['producto']
        if producto not in total['productos']:
            total['productos'][producto] = '='

        total['productos'][producto] += subtotal_row


    sheet.append(['TOTAL', None, None, None, None, None, None, None,
                  total['origen'],
                  total['destino'],
                  total['diferencia'],
                  total['tolerancia'],
                  total['diferencia_tolerancia'],
                  None, total['total']])

    # Obtener el rango de columnas con valores None
    inicio_columna = 1  # Cambiar al índice de la primera columna con valor None
    fin_columna = 8   # Cambiar al índice de la última columna con valor None

    # Combinar las celdas en el rango de columnas
    sheet.merge_cells(start_row=sheet.max_row, start_column=inicio_columna,
                      end_row=sheet.max_row, end_column=fin_columna)

    # Centrar el contenido en la celda combinada
    merged_cell = sheet.cell(row=sheet.max_row, column=inicio_columna)
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Formatear columnas 8 a 15 como números
    for col in range(8, 16):
        cell = sheet.cell(row=sheet.max_row, column=col)

        if col == 14:
            cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED2
        else:
            cell.number_format = "#,##0"

    for col in range(1, 16):
        cell = sheet.cell(row=sheet.max_row, column=col)
        thin_border = Border(left=Side(style='thin'), right=Side(
            style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.border = thin_border

        fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        cell.fill = fill

    sheet.append([None, None, None, None, None, None, None, None,
                None,  None, None, None, None, None, f'=+O{last_row + 1}/11'])
    cell = sheet.cell(row=sheet.max_row, column=15)
    cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED2

    sheet.append([])
    for producto, subtotal in total['productos'].items():
        sheet.append([None, None, None, None, None, None, None, None,
                None,  None, None, None, None, producto, subtotal])
        cell = sheet.cell(row=sheet.max_row, column=15)
        cell.number_format = "#,##0"

    # Guardar el archivo Excel en el flujo de salida
    workbook.save(output)
    output.seek(0)

    # Crear la respuesta para el cliente con el archivo Excel
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers[
        'Content-Disposition'] = f'attachment; filename=cobranza_{fecha_creacion}.xlsx'
    logger.warning(f'Cobranza exportada {fecha_creacion}')

    return response


def exportar_informe_planillas(fecha_inicio, fecha_fin):
    fecha_inicio_parsed = parser.isoparse(fecha_inicio).date()
    fecha_fin_parsed = parser.isoparse(fecha_fin).date()

    # Query the database for records within the date range
    cobranzas = Cobranzas.query.filter(Cobranzas.fecha_creacion.between(
        fecha_inicio_parsed, fecha_fin_parsed)).all()

    # Ordena las cobranzas por las columnas 'origen' y 'destino'
    cobranzas_ordenadas = sorted(
        cobranzas, key=lambda cobranza: cobranza.chapa)

    # Crear un archivo Excel en memoria
    output = io.BytesIO()
    workbook = Workbook()
    sheet = workbook.active

    sheet.column_dimensions['A'].width = 2.64
    sheet.column_dimensions['B'].width = 11.00
    sheet.column_dimensions['C'].width = 20.55
    sheet.column_dimensions['D'].width = 9.09
    sheet.column_dimensions['E'].width = 11.82
    sheet.column_dimensions['F'].width = 18.64
    sheet.column_dimensions['G'].width = 17.64
    sheet.column_dimensions['H'].width = 9.91
    sheet.column_dimensions['I'].width = 10.91
    sheet.column_dimensions['J'].width = 11.09
    sheet.column_dimensions['K'].width = 7.18
    sheet.column_dimensions['L'].width = 6.27
    sheet.column_dimensions['M'].width = 6.36
    sheet.column_dimensions['N'].width = 6.27
    sheet.column_dimensions['O'].width = 14.64

    # Agregar encabezados
    encabezados = ['N°', 'Fecha', 'Chofer', 'Chapa', 'Producto', 'Origen', 'Destino', 'Tiquet',
                   'Kilos Origen', 'Kilos Destino', 'Dif.', 'Tolera', 'Dif. Tol.', 'Precio', 'Total']
    sheet.append(encabezados)

    # Congelar la fila del encabezado
    sheet.freeze_panes = 'A2'

    # Agregar filas de datos
    contador = 2
    for cobranza in cobranzas_ordenadas:
        fila = [contador - 1,
                cobranza.fecha_viaje.strftime('%d/%m/%Y'),
                cobranza.chofer,
                cobranza.chapa,
                cobranza.producto,
                cobranza.origen,
                cobranza.destino,
                cobranza.tiquet,
                cobranza.kilos_origen,
                cobranza.kilos_destino,
                f'=+J{contador}-I{contador}',
                f'=ROUND(J{contador}*0.002, 0)',
                f'=+L{contador}+K{contador}',
                cobranza.precio,
                f'=ROUND(J{contador}*N{contador}, 0)']

        sheet.append(fila)

        for col in range(8, 16):
            cell = sheet.cell(row=sheet.max_row, column=col)

            if col == 14:
                cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED2
            else:
                cell.number_format = "#,##0"

        contador += 1

    # Guardar el archivo Excel en el flujo de salida
    workbook.save(output)
    output.seek(0)

    # Crear la respuesta para el cliente con el archivo Excel
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers[
        'Content-Disposition'] = f'attachment; filename=informe_{fecha_inicio}-{fecha_fin}.xlsx'
    logger.warning(f'Informe exportado: {fecha_inicio}-{fecha_fin}')

    return response


def exportar_liquidacion(chofer, fecha):
    liq_id = Liquidaciones.query.filter_by(chofer=chofer, fecha_liquidacion=fecha).first().id

    # Perform an inner join between Cobranza and Liquidaciones based on 'chofer' and 'fecha_de_liquidacion'
    viajes = db.session.query(Cobranzas, LiquidacionViajes).join(
        LiquidacionViajes,
        Cobranzas.id == LiquidacionViajes.id
    ).filter(
        Cobranzas.chofer == chofer,
        LiquidacionViajes.id_liquidacion == liq_id
    ).order_by(Cobranzas.fecha_viaje).all()

    # Perform an inner join between Cobranza and Liquidaciones based on 'chofer' and 'fecha_de_liquidacion'
    gastos_sin_boleta = LiquidacionGastos.query.filter_by(
        id_liquidacion = liq_id, boleta=None).all()
    gastos_con_boleta = LiquidacionGastos.query.filter_by(
        id_liquidacion = liq_id).filter(LiquidacionGastos.boleta.isnot(None)).all()

    # Calcular la longitud máxima de las tres listas
    max_len = max(len(viajes), len(gastos_sin_boleta), len(gastos_con_boleta))

    # Rellenar las listas para que tengan la misma longitud con None si es necesario
    viajes += [None] * (max_len - len(viajes))
    gastos_sin_boleta += [None] * (max_len - len(gastos_sin_boleta))
    gastos_con_boleta += [None] * (max_len - len(gastos_con_boleta))

    # Combinar las tres listas en una lista de tuplas usando zip
    results = zip(viajes, gastos_sin_boleta, gastos_con_boleta)

    # Crear un archivo Excel en memoria
    output = io.BytesIO()
    workbook = Workbook()
    sheet = workbook.active

    sheet.column_dimensions['A'].width = 2.64
    sheet.column_dimensions['B'].width = 9.91
    sheet.column_dimensions['C'].width = 7.91
    sheet.column_dimensions['D'].width = 9.36
    sheet.column_dimensions['E'].width = 18
    sheet.column_dimensions['F'].width = 18
    sheet.column_dimensions['G'].width = 10.3
    sheet.column_dimensions['H'].width = 11.0
    sheet.column_dimensions['I'].width = 8.09
    sheet.column_dimensions['J'].width = 9.60
    sheet.column_dimensions['K'].width = 10.27
    sheet.column_dimensions['L'].width = 10.82
    sheet.column_dimensions['M'].width = 10.27
    sheet.column_dimensions['N'].width = 10.36
    sheet.column_dimensions['O'].width = 8.91
    sheet.column_dimensions['P'].width = 10.82

    # Agregar la fecha como la primera fila
    sheet.append([])  # Agregar una fila en blanco después de la fecha
    # Agregar una fila en blanco después de la fecha
    sheet.append(['LIQUIDACION DE FLETES'])

    # Obtener el rango de columnas con valores None
    inicio_columna = 1  # Cambiar al índice de la primera columna con valor None
    fin_columna = 16   # Cambiar al índice de la última columna con valor None

    # Combinar las celdas en el rango de columnas
    sheet.merge_cells(start_row=sheet.max_row, start_column=inicio_columna,
                      end_row=sheet.max_row, end_column=fin_columna)

    # Centrar el contenido en la celda combinada
    merged_cell = sheet.cell(row=sheet.max_row, column=inicio_columna)
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Aplicar bordes
    thin_border = Border(left=Side(style='thin'), right=Side(
        style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for col in range(1, 17):
        cell = sheet.cell(row=sheet.max_row, column=col)
        thin_border = Border(left=Side(style='thin'), right=Side(
            style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.border = thin_border
        cell.font = Font(bold=True)

    sheet.row_dimensions[2].height = 21

    if viajes and viajes[0]:
        chapa = viajes[0][0].chapa
    else:
        chapa = ""

    # Agregar una fila en blanco después de la fecha
    sheet.append(
        [f'Conductor: {chofer}                Chapa: {chapa}                Fecha: {datetime.now().strftime("%d/%m/%Y")}'])
    # Obtener el rango de columnas con valores None
    inicio_columna = 1  # Cambiar al índice de la primera columna con valor None
    fin_columna = 16   # Cambiar al índice de la última columna con valor None

    # Combinar las celdas en el rango de columnas
    sheet.merge_cells(start_row=sheet.max_row, start_column=inicio_columna,
                      end_row=sheet.max_row, end_column=fin_columna)

    # Centrar el contenido en la celda combinada
    merged_cell = sheet.cell(row=sheet.max_row, column=inicio_columna)
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')

    thin_border = Border(left=Side(style='thin'), right=Side(
        style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for col in range(1, 17):
        cell = sheet.cell(row=sheet.max_row, column=col)
        thin_border = Border(left=Side(style='thin'), right=Side(
            style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.border = thin_border
        cell.font = Font(bold=True)

    sheet.row_dimensions[3].height = 21

    sheet.append(['FLETES', None, None, None, None, None, None, None, None, None, None,
                  'GASTOS (VIATICO/GASOIL)', None, None, None, None])

    # Obtener el rango de columnas con valores None
    inicio_columna = 1  # Cambiar al índice de la primera columna con valor None
    fin_columna = 11   # Cambiar al índice de la última columna con valor None

    # Combinar las celdas en el rango de columnas
    sheet.merge_cells(start_row=sheet.max_row, start_column=inicio_columna,
                      end_row=sheet.max_row, end_column=fin_columna)

    # Centrar el contenido en la celda combinada
    merged_cell = sheet.cell(row=sheet.max_row, column=inicio_columna)
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')

    thin_border = Border(left=Side(style='thin'), right=Side(
        style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for col in range(1, 17):
        cell = sheet.cell(row=sheet.max_row, column=col)
        thin_border = Border(left=Side(style='thin'), right=Side(
            style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.border = thin_border
        cell.font = Font(bold=True)

    # Obtener el rango de columnas con valores None
    inicio_columna = 12  # Cambiar al índice de la primera columna con valor None
    fin_columna = 16   # Cambiar al índice de la última columna con valor None

    # Combinar las celdas en el rango de columnas
    sheet.merge_cells(start_row=sheet.max_row, start_column=inicio_columna,
                      end_row=sheet.max_row, end_column=fin_columna)

    # Centrar el contenido en la celda combinada
    merged_cell = sheet.cell(row=sheet.max_row, column=inicio_columna)
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')

    thin_border = Border(left=Side(style='thin'), right=Side(
        style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for col in range(1, 17):
        cell = sheet.cell(row=sheet.max_row, column=col)
        thin_border = Border(left=Side(style='thin'), right=Side(
            style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.border = thin_border
        cell.font = Font(bold=True)

    # Agregar encabezados
    encabezados = ['N°', 'Fecha', 'Producto', 'Recepcion N°', 'Origen', 'Destino', 'Kilos Origen', 'Kilos Llegada',
                   'Dif.', 'Gs. p/ KILO', 'Importe Gs.', 'Fecha', 'Importe Gs.', 'Fecha', 'Boleta N°', 'Importe Gs.']
    sheet.append(encabezados)

    for col in range(1, 17):
        cell = sheet.cell(row=sheet.max_row, column=col)
        thin_border = Border(left=Side(style='thin'), right=Side(
            style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.border = thin_border
        cell.font = Font(bold=True)

    sheet.row_dimensions[5].height = 25

    contador = 1
    for viaje, sin_boleta, con_boleta in results:
        fila = [contador]

        if viaje:
            current_row = contador + 5
            diferencia = f'=+H{current_row}-G{current_row}'
            total_gs = f'=ROUND(H{current_row}*J{current_row}, 0)'

            viaje_fila = [viaje[0].fecha_viaje.strftime('%d/%m/%Y'), viaje[0].producto,
                          viaje[0].tiquet, viaje[0].origen, viaje[0].destino,
                          viaje[0].kilos_origen, viaje[0].kilos_destino,
                          diferencia, 
                          viaje[1].precio_liquidacion,
                          total_gs]
            fila.extend(viaje_fila)

        else:
            fila.extend([None] * 10)

        if sin_boleta:
            sin_boleta_fila = [sin_boleta.fecha.strftime(
                '%d/%m/%Y'), sin_boleta.importe]
            fila.extend(sin_boleta_fila)
        else:
            fila.extend([None] * 2)

        if con_boleta:
            con_boleta_fila = [con_boleta.fecha.strftime(
                '%d/%m/%Y'), con_boleta.boleta, con_boleta.importe]
            fila.extend(con_boleta_fila)
        else:
            fila.extend([None] * 3)

        sheet.append(fila)

        for col in range(1, 17):
            cell = sheet.cell(row=sheet.max_row, column=col)

            if col == 10:
                cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED2
            else:
                cell.number_format = "#,##0"

            thin_border = Border(left=Side(style='thin'), right=Side(
                style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            cell.border = thin_border

        contador += 1

    last_row = 5 + max_len
    subtotal_sin_boleta = f'=SUM(M6:M{last_row})'
    subtotal_con_boleta = f'=SUM(P6:P{last_row})'
    subtotales = [None, None, None, None, None, None, None, None, None, None,
                  None, 'Subtotal', subtotal_sin_boleta, None, 'Subtotal', subtotal_con_boleta]
    sheet.append(subtotales)

    for col in range(1, 17):
        cell = sheet.cell(row=sheet.max_row, column=col)
        thin_border = Border(left=Side(style='thin'), right=Side(
            style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.border = thin_border
        cell.font = Font(bold=True)

    for col in [13, 16]:
        cell = sheet.cell(row=sheet.max_row, column=col)
        cell.number_format = "#,##0"
        thin_border = Border(left=Side(style='thin'), right=Side(
            style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.border = thin_border

    total_gastos = f'=+M{last_row + 1}+P{last_row + 1}'
    subtotal_viajes = f'=SUM(K6:K{last_row})'
    total = [None, None, None, None, None, None, None, None, 'TOTAL FLETES:',
             None, subtotal_viajes, 'TOTAL GASTOS:', None, None, None, total_gastos]
    sheet.append(total)

    for col in [11, 16]:
        cell = sheet.cell(row=sheet.max_row, column=col)
        cell.number_format = "#,##0"
        thin_border = Border(left=Side(style='thin'), right=Side(
            style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.border = thin_border

    for col in range(1, 17):
        cell = sheet.cell(row=sheet.max_row, column=col)
        thin_border = Border(left=Side(style='thin'), right=Side(
            style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.border = thin_border
        cell.font = Font(bold=True)

    sheet.append([None])

    total_cobrar = [None, None, None, None, None, None, None,
                    'TOTAL A COBRAR:', None, None, f'=+K{last_row + 2}-P{last_row + 2}']
    sheet.append(total_cobrar)

    for col in range(8, 12):
        cell = sheet.cell(row=sheet.max_row, column=col)
        thin_border = Border(left=Side(style='thin'), right=Side(
            style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.border = thin_border
        cell.font = Font(bold=True)

        if col == 11:
            cell.number_format = "#,##0"

    total_facturar = [None, None, None, None, None, None, None,
                      'TOTAL A FACTURAR:', None, None, f'=+K{last_row + 2}-P{last_row + 1}']
    sheet.append(total_facturar)

    for col in range(8, 12):
        cell = sheet.cell(row=sheet.max_row, column=col)
        thin_border = Border(left=Side(style='thin'), right=Side(
            style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.border = thin_border
        cell.font = Font(bold=True)

        if col == 11:
            cell.number_format = "#,##0"

    sheet.append([None, None, None, None, None, None, None,
                      'Facturar a nombre de CARMELO MEDINA. Ruc: 850.299-4', None, None, None])
    
    sheet.merge_cells(start_row=sheet.max_row, start_column=8,
                      end_row=sheet.max_row, end_column=12)

    for col in range(8, 13):
        cell = sheet.cell(row=sheet.max_row, column=col)
        thin_border = Border(left=Side(style='thin'), right=Side(
            style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.border = thin_border


    sheet.append([])

    sheet.append([None, None, None, None, None, None, None,
                    'Descripcion', None, 'Exenta', 'IVA 5%', 'IVA 10%'])
    
    sheet.merge_cells(start_row=sheet.max_row, start_column=8,
                    end_row=sheet.max_row, end_column=9)

    for col in range(8, 13):
        cell = sheet.cell(row=sheet.max_row, column=col)
        thin_border = Border(left=Side(style='thin'), right=Side(
            style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.border = thin_border

    
    sheet.append([None, None, None, None, None, None, None,
                'Servicio de Flete', None, 0, 0, f'=+K{last_row + 5}'])
    
    sheet.merge_cells(start_row=sheet.max_row, start_column=8,
                    end_row=sheet.max_row, end_column=9)

    for col in range(8, 13):
        cell = sheet.cell(row=sheet.max_row, column=col)
        thin_border = Border(left=Side(style='thin'), right=Side(
            style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.border = thin_border

        if col >= 10:
            cell.number_format = "#,##0"

    sheet.append([None, None, None, None, None, None, None,
                None, None, None, None])
    
    sheet.merge_cells(start_row=sheet.max_row, start_column=8,
                end_row=sheet.max_row, end_column=9)

    for col in range(8, 13):
        cell = sheet.cell(row=sheet.max_row, column=col)
        thin_border = Border(left=Side(style='thin'), right=Side(
            style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.border = thin_border

        if col >= 10:
            cell.number_format = "#,##0"


    sheet.append([None, None, None, None, None, None, None,
                'Subtotal', None, f'=+J{last_row + 9}', 0, f'=+L{last_row + 9}'])
    
    sheet.merge_cells(start_row=sheet.max_row, start_column=8,
                end_row=sheet.max_row, end_column=9)
    
    for col in range(8, 13):
        cell = sheet.cell(row=sheet.max_row, column=col)
        thin_border = Border(left=Side(style='thin'), right=Side(
            style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.border = thin_border

        if col >= 10:
            cell.number_format = "#,##0"


    sheet.append([None, None, None, None, None, None, None,
            'Total', None, None, None, f'=+J{last_row + 11}+L{last_row + 11}'])
    
    sheet.merge_cells(start_row=sheet.max_row, start_column=8,
                end_row=sheet.max_row, end_column=9)
    
    for col in range(8, 13):
        cell = sheet.cell(row=sheet.max_row, column=col)
        thin_border = Border(left=Side(style='thin'), right=Side(
            style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.border = thin_border
        cell.font = Font(bold=True)

        if col >= 10:
            cell.number_format = "#,##0"

    sheet.append([None, None, None, None, None, None, None, None,
        None, 'IVA 10%', None, f'=+L{last_row + 12}/11'])
    
    sheet.merge_cells(start_row=sheet.max_row, start_column=8,
                end_row=sheet.max_row, end_column=9)

    for col in range(8, 13):
        cell = sheet.cell(row=sheet.max_row, column=col)
        thin_border = Border(left=Side(style='thin'), right=Side(
            style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.border = thin_border
        cell.font = Font(bold=True)

        if col >= 10:
            cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED2
    
    # Guardar el archivo Excel en el flujo de salida
    workbook.save(output)
    output.seek(0)

    # Crear la respuesta para el cliente con el archivo Excel
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={chofer}_Liquidacion_{fecha}.xlsx'
    logger.warning(f'Liquidacion {chofer}_{fecha} exportada')

    return response


def exportar_precios():
    precios = Precios.query.all()

    # Ordena las cobranzas por las columnas 'origen' y 'destino'
    precios_ordenados = sorted(precios, key=lambda precios: (precios.origen, precios.destino))

    # Crear un archivo Excel en memoria
    output = io.BytesIO()
    workbook = Workbook()
    sheet = workbook.active

    # Agregar encabezados
    encabezados = ['Origen', 'Destino', 'Precio', 'Precio de Liquidación']
    sheet.append(encabezados)

    # Ajustar el ancho de las columnas
    for col_idx in range(1, 5):
        column_letter = get_column_letter(col_idx)
        sheet.column_dimensions[column_letter].width = 20

    # Estilo de borde
    border_style = Border(left=Side(style='thin'),
                          right=Side(style='thin'),
                          top=Side(style='thin'),
                          bottom=Side(style='thin'))

    # Aplicar el estilo de borde a cada celda en la fila
    for cell in sheet[sheet.max_row]:
        cell.border = border_style

    # Agregar filas de datos
    for precio in precios_ordenados:
        fila = [precio.origen, precio.destino,
                precio.precio, precio.precio_liquidacion]

        sheet.append(fila)

        sheet.cell(row=sheet.max_row,
                   column=3).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
        sheet.cell(row=sheet.max_row,
                   column=4).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

        # Aplicar el estilo de borde a cada celda en la fila
        for cell in sheet[sheet.max_row]:
            cell.border = border_style

    # Guardar el archivo Excel en el flujo de salida
    workbook.save(output)
    output.seek(0)

    # Crear la respuesta para el cliente con el archivo Excel
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=lista_de_precios.xlsx'

    logger.warning("Lista de Precios exportada")
    return response
